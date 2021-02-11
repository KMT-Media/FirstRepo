# barchart
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# load workbook - return workbook object
wb = xl.load_workbook('file/Barchart.xlsx')

# access sheet
sheet = wb['Sheet1']

# access cell
# cell = sheet.cell(1, 1)

# how many rows we have
# print(sheet.max_row)

# iterate through each rows in excel file
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_cell = sheet.cell(row, 4)
    corrected_cell.value = corrected_price

# wb.save('file/Barchart2.xlsx')